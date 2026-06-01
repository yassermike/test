"""
Générateur de rapports - Export en CSV, Excel et HTML
"""
import pandas as pd
import io
from typing import List, Dict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


class ReportGenerator:
    """Génère des rapports d'analyse à partir des erreurs détectées"""
    
    def __init__(self, df: pd.DataFrame, errors: List[Dict], structure: str = ""):
        self.df = df
        self.errors = errors
        self.structure = structure
        self.errors_df = pd.DataFrame(errors) if errors else pd.DataFrame()
    
    def to_csv(self) -> bytes:
        """Export CSV des erreurs"""
        if self.errors_df.empty:
            return "Aucune erreur détectée".encode('utf-8')
        return self.errors_df.to_csv(index=False).encode('utf-8-sig')  # BOM pour Excel
    
    def to_excel_with_corrections(self) -> bytes:
        """Génère un Excel avec:
        - Feuille 1: données originales avec cellules colorées
        - Feuille 2: liste des erreurs
        - Feuille 3: synthèse
        """
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Feuille 1: Données originales
            self.df.to_excel(writer, sheet_name='Données', index=False)
            
            # Feuille 2: Erreurs détaillées
            if not self.errors_df.empty:
                self.errors_df.to_excel(writer, sheet_name='Erreurs', index=False)
            
            # Feuille 3: Synthèse
            synthesis = self._build_synthesis()
            synthesis.to_excel(writer, sheet_name='Synthèse', index=False)
            
            # Mise en forme
            workbook = writer.book
            
            # Colorer les cellules erronées dans "Données"
            if not self.errors_df.empty and 'ligne' in self.errors_df.columns:
                ws = workbook['Données']
                
                # Couleurs par sévérité
                colors = {
                    'critique': PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid'),
                    'moderee': PatternFill(start_color='FFE5CC', end_color='FFE5CC', fill_type='solid'),
                    'mineure': PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid'),
                }
                
                for _, error in self.errors_df.iterrows():
                    try:
                        ligne = int(error.get('ligne', 0))
                        colonne = error.get('colonne', '')
                        severite = error.get('severite', 'moderee')
                        
                        if ligne > 0 and colonne in self.df.columns:
                            col_idx = list(self.df.columns).index(colonne) + 1
                            cell = ws.cell(row=ligne + 1, column=col_idx)  # +1 car header
                            cell.fill = colors.get(severite, colors['moderee'])
                    except (ValueError, KeyError):
                        continue
                
                # En-têtes en gras
                for cell in ws[1]:
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='1F77B4', end_color='1F77B4', fill_type='solid')
            
            # Mise en forme de la feuille Erreurs
            if 'Erreurs' in workbook.sheetnames:
                ws_err = workbook['Erreurs']
                for cell in ws_err[1]:
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='D62728', end_color='D62728', fill_type='solid')
                
                # Auto-fit columns
                for col_cells in ws_err.columns:
                    max_length = max(len(str(cell.value or "")) for cell in col_cells)
                    ws_err.column_dimensions[col_cells[0].column_letter].width = min(max_length + 2, 50)
            
            # Mise en forme synthèse
            ws_syn = workbook['Synthèse']
            for cell in ws_syn[1]:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='2CA02C', end_color='2CA02C', fill_type='solid')
        
        output.seek(0)
        return output.getvalue()
    
    def _build_synthesis(self) -> pd.DataFrame:
        """Construit la synthèse"""
        synthesis_data = []
        
        synthesis_data.append({"Indicateur": "Date du rapport", "Valeur": datetime.now().strftime("%Y-%m-%d %H:%M:%S")})
        synthesis_data.append({"Indicateur": "Lignes analysées", "Valeur": len(self.df)})
        synthesis_data.append({"Indicateur": "Colonnes", "Valeur": len(self.df.columns)})
        synthesis_data.append({"Indicateur": "Total anomalies", "Valeur": len(self.errors)})
        
        if not self.errors_df.empty:
            if 'severite' in self.errors_df.columns:
                for sev in ['critique', 'moderee', 'mineure']:
                    count = (self.errors_df['severite'] == sev).sum()
                    synthesis_data.append({"Indicateur": f"Anomalies {sev}", "Valeur": int(count)})
            
            if 'type_erreur' in self.errors_df.columns:
                synthesis_data.append({"Indicateur": "---", "Valeur": "---"})
                synthesis_data.append({"Indicateur": "RÉPARTITION PAR TYPE", "Valeur": ""})
                type_counts = self.errors_df['type_erreur'].value_counts()
                for type_err, count in type_counts.items():
                    synthesis_data.append({"Indicateur": f"  {type_err}", "Valeur": int(count)})
            
            if 'colonne' in self.errors_df.columns:
                synthesis_data.append({"Indicateur": "---", "Valeur": "---"})
                synthesis_data.append({"Indicateur": "TOP COLONNES PROBLÉMATIQUES", "Valeur": ""})
                col_counts = self.errors_df['colonne'].value_counts().head(5)
                for col, count in col_counts.items():
                    synthesis_data.append({"Indicateur": f"  {col}", "Valeur": int(count)})
            
            # Taux de qualité
            total_cells = len(self.df) * len(self.df.columns)
            quality_rate = (1 - len(self.errors) / total_cells) * 100 if total_cells > 0 else 100
            synthesis_data.append({"Indicateur": "---", "Valeur": "---"})
            synthesis_data.append({"Indicateur": "TAUX DE QUALITÉ", "Valeur": f"{quality_rate:.2f}%"})
        
        return pd.DataFrame(synthesis_data)
    
    def to_html_report(self) -> bytes:
        """Génère un rapport HTML stylé"""
        
        # Calculs
        total_errors = len(self.errors)
        critical_count = sum(1 for e in self.errors if e.get('severite') == 'critique')
        moderate_count = sum(1 for e in self.errors if e.get('severite') == 'moderee')
        minor_count = sum(1 for e in self.errors if e.get('severite') == 'mineure')
        
        # Tableau des erreurs
        errors_html = ""
        if not self.errors_df.empty:
            errors_html = self.errors_df.to_html(classes='errors-table', index=False, escape=False)
        
        html = f"""<!DOCTYPE html>
<html lang="fr">
<head>
    <meta charset="UTF-8">
    <title>Rapport de Validation - Études Client Mystère</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: 'Segoe UI', Arial, sans-serif;
            background: #f5f7fa;
            color: #333;
            padding: 20px;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            border-radius: 12px;
            box-shadow: 0 4px 20px rgba(0,0,0,0.08);
            overflow: hidden;
        }}
        header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 40px;
            text-align: center;
        }}
        header h1 {{ font-size: 2.5em; margin-bottom: 10px; }}
        header p {{ opacity: 0.9; }}
        .content {{ padding: 40px; }}
        .stats {{
            display: grid;
            grid-template-columns: repeat(4, 1fr);
            gap: 20px;
            margin-bottom: 40px;
        }}
        .stat-card {{
            padding: 25px;
            border-radius: 10px;
            text-align: center;
            color: white;
        }}
        .stat-card h2 {{ font-size: 3em; margin-bottom: 5px; }}
        .stat-card p {{ font-size: 1em; opacity: 0.95; }}
        .stat-total {{ background: linear-gradient(135deg, #667eea, #764ba2); }}
        .stat-critical {{ background: linear-gradient(135deg, #ee0979, #ff6a00); }}
        .stat-moderate {{ background: linear-gradient(135deg, #f7971e, #ffd200); color: #333; }}
        .stat-minor {{ background: linear-gradient(135deg, #56ab2f, #a8e063); }}
        h2.section-title {{
            color: #2c3e50;
            margin: 30px 0 20px;
            padding-bottom: 10px;
            border-bottom: 3px solid #667eea;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
            font-size: 0.9em;
        }}
        table th {{
            background: #2c3e50;
            color: white;
            padding: 12px;
            text-align: left;
        }}
        table td {{
            padding: 10px 12px;
            border-bottom: 1px solid #ecf0f1;
        }}
        table tr:nth-child(even) {{ background: #f8f9fa; }}
        table tr:hover {{ background: #e3f2fd; }}
        .footer {{
            background: #2c3e50;
            color: white;
            text-align: center;
            padding: 20px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <header>
            <h1>📊 Rapport de Validation</h1>
            <p>Études Client Mystère & Satisfaction</p>
            <p style="margin-top:10px;font-size:0.9em;">Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}</p>
        </header>
        
        <div class="content">
            <div class="stats">
                <div class="stat-card stat-total">
                    <h2>{total_errors}</h2>
                    <p>Total anomalies</p>
                </div>
                <div class="stat-card stat-critical">
                    <h2>{critical_count}</h2>
                    <p>🔴 Critiques</p>
                </div>
                <div class="stat-card stat-moderate">
                    <h2>{moderate_count}</h2>
                    <p>🟠 Modérées</p>
                </div>
                <div class="stat-card stat-minor">
                    <h2>{minor_count}</h2>
                    <p>🟡 Mineures</p>
                </div>
            </div>
            
            <h2 class="section-title">📋 Informations générales</h2>
            <ul style="line-height: 2; list-style: none;">
                <li><strong>Lignes analysées:</strong> {len(self.df)}</li>
                <li><strong>Colonnes:</strong> {len(self.df.columns)}</li>
                <li><strong>Taux de qualité:</strong> {(1 - total_errors / (len(self.df) * len(self.df.columns))) * 100:.2f}%</li>
            </ul>
            
            <h2 class="section-title">⚠️ Détail des anomalies</h2>
            {errors_html if errors_html else '<p style="text-align:center;color:#27ae60;font-size:1.2em;">✅ Aucune anomalie détectée!</p>'}
        </div>
        
        <div class="footer">
            <p>Rapport généré par la Plateforme de Validation - Propulsé par Claude AI</p>
        </div>
    </div>
</body>
</html>"""
        
        return html.encode('utf-8')
