"""
Validateur spécifique Client Mystère SGA
Règles métier codées en dur - 7 règles de vérification
"""
import pandas as pd
import numpy as np
import re
import io
from typing import List, Dict, Tuple, Optional
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# UTILITAIRES
# ============================================================

def normalize(s: str) -> str:
    """Normalise une chaîne pour la comparaison (minuscules, sans accents)"""
    if not isinstance(s, str):
        return ''
    s = s.lower().strip()
    for a, b in [('é','e'),('è','e'),('ê','e'),('ë','e'),('à','a'),('â','a'),
                 ('î','i'),('ï','i'),('ô','o'),('ù','u'),('û','u'),('ü','u'),('ç','c')]:
        s = s.replace(a, b)
    return s


def find_col(df: pd.DataFrame, keywords: List[str]) -> Optional[str]:
    """Trouve la première colonne contenant tous les mots-clés (insensible accents/casse)"""
    for col in df.columns:
        cn = normalize(col)
        if all(normalize(k) in cn for k in keywords):
            return col
    return None


def find_col_any(df: pd.DataFrame, keyword_groups: List[List[str]]) -> Optional[str]:
    """Retourne la première colonne correspondant à l'un des groupes de mots-clés"""
    for kw_list in keyword_groups:
        col = find_col(df, kw_list)
        if col:
            return col
    return None


WILAYAS = [
    'MSILA', 'AIN TIMOUCHENT', 'ALGER', 'ANNABA', 'BATNA', 'BBA',
    'BEJAIA', 'BÉJAIA', 'BISKRA', 'BLIDA', 'BOUIRA', 'BOUMERDES', 'CHLEF',
    'CONSTANTINE', 'ELTAREF', 'GHARDAIA', 'JIJEL', 'KHENCHELA', 'MASCARA',
    'MEDEA', 'MILA', 'MOSTAGANEM', 'ORAN', 'OUARGLA', 'RELIZANE', 'SETIF',
    'SIDI BEL ABBES', 'SKIKDA', 'SOUK-AHRAS', 'TIARET', 'TIPAZA',
    'TIZI OUZOU', 'TLEMCEN', 'OUM EL BOUAGHI', 'AIN DEFLA'
]


def get_city_cols(df: pd.DataFrame) -> List[str]:
    """Retourne les colonnes représentant les adresses d'agence par wilaya"""
    result = []
    for col in df.columns:
        if '[Other' in col or 'maps.google' in col:
            continue
        cu = col.strip().upper().replace('É','E').replace('È','E').replace('Ê','E').replace('Â','A').replace('Ï','I')
        for w in WILAYAS:
            wu = w.upper()
            if cu == wu or cu == ' ' + wu:
                result.append(col)
                break
    return result


def get_agence_id(row: pd.Series, city_cols: List[str]) -> str:
    """Extrait l'identifiant unique de l'agence depuis une ligne"""
    for col in city_cols:
        val = row.get(col)
        if pd.notna(val) and str(val).strip() and str(val) != 'nan':
            return str(val).strip()
    return 'AGENCE_INCONNUE'


def parse_time_slot(slot) -> Optional[Tuple[int, int, int, int]]:
    """Parse '13h315h30' ou '13h30-15h30' → (h_start, m_start, h_end, m_end)"""
    if pd.isna(slot):
        return None
    s = str(slot).replace(' ', '')
    match = re.search(r'(\d{1,2})[hH](\d{2})[-–]?(\d{1,2})[hH](\d{2})', s)
    if match:
        return int(match.group(1)), int(match.group(2)), int(match.group(3)), int(match.group(4))
    return None


def parse_minute_slot(slot) -> Optional[Tuple[int, int]]:
    """Parse '0 - 5 MINUTES' → (0, 5), '>20 MINUTES' → (21, 999)"""
    if pd.isna(slot):
        return None
    s = str(slot).upper()
    if '>20' in s or '> 20' in s:
        return (21, 999)
    match = re.search(r'(\d+)\s*[-–]\s*(\d+)', s)
    if match:
        return int(match.group(1)), int(match.group(2))
    return None


def sat_to_score(val) -> int:
    if pd.isna(val):
        return -1
    v = normalize(str(val))
    if 'tres satisfait' in v or 'très satisfait' in v:
        return 5
    if 'satisfait' in v and 'in' not in v:
        return 4
    if 'ni satisfait' in v or 'neutre' in v or 'moyen' in v:
        return 3
    if 'tres insatisfait' in v or 'très insatisfait' in v:
        return 1
    if 'insatisfait' in v:
        return 2
    return -1


def reco_to_score(val) -> int:
    if pd.isna(val):
        return -1
    v = normalize(str(val))
    if 'certainement' in v or 'fortement' in v or 'absolument' in v:
        return 5
    if 'recommanderai' in v and 'pas' not in v and 'ne ' not in v:
        return 4
    if 'hesite' in v or 'peut-etre' in v or 'eventuellement' in v or 'probablement' in v:
        return 3
    if ('pas' in v and 'recommand' in v) or ('ne ' in v and 'recommand' in v):
        return 2
    if 'jamais' in v or 'deconseill' in v:
        return 1
    return -1


# ============================================================
# MOTEUR DE VALIDATION
# ============================================================

class SGAValidator:

    RULES_META = {
        'R1': ('Couverture agence (3 visites)', 'critique'),
        'R2': ('Mix profils PRI/PRO par agence', 'critique'),
        'R3': ('Diversité scénarios par agence', 'moderee'),
        'R4': ('Cohérence note/satisfaction/recommandation', 'moderee'),
        'R5': ('Cohérence heure visite / créneau', 'moderee'),
        'R6': ('Cohérence temps attente / créneau', 'mineure'),
        'R7': ('Cohérence temps conseiller / créneau', 'mineure'),
    }

    def __init__(self, df: pd.DataFrame):
        self.df = df.reset_index(drop=True).copy()
        self.errors: List[Dict] = []

        # Colonnes identité
        self.city_cols = get_city_cols(df)
        self.col_sbjnum   = 'SbjNum' if 'SbjNum' in df.columns else df.columns[0]
        self.col_srvyr    = find_col_any(df, [['Srvyr'], ['srvyr'], ['enqueteur']])
        self.col_wilaya   = find_col(df, ['WILAYA'])
        self.col_banque   = find_col(df, ['Banque'])

        # Colonnes questionnaire clés
        self.col_profil         = find_col_any(df, [['PROFIL DU CLIENT'], ['profil du client'], ['Q3']])
        self.col_scen_pri       = find_col(df, ['SCENARIOS PRI'])
        self.col_scen_pro       = find_col_any(df, [['SCENARIOS PR0'], ['SCENARIOS PRO']])
        self.col_scen_corpo     = find_col(df, ['SCENARIOS CORPO'])
        self.col_heure          = find_col_any(df, [['heure de visite'], ['enregistrer l', 'heure']])
        self.col_heure_grille   = find_col_any(df, [['heure dans la grille'], ['BIS Q2', 'grille']])
        self.col_att_min        = find_col_any(df, [['combien de temps avez-vous attendu'], ['TEMPS D', 'ATTENTE', 'Q23']])
        self.col_att_grille     = find_col_any(df, [['RECORDER', 'TEMPS D', 'ATTENTE', 'GRILLE'], ['ATTENTE CETTE GRILLE']])
        self.col_cons_min       = find_col_any(df, [['temps passe avec le guichetier'], ['temps avez-vous passe avec']])
        self.col_cons_grille    = find_col_any(df, [['RECORDER', 'TEMPS PASSE', 'GRILLE'], ['TEMPS PASSE', 'GUICHETIER', 'GRILLE']])
        self.col_note           = find_col_any(df, [['valoris'], ['Q1', 'visite', 'client']])
        self.col_satisfaction   = find_col_any(df, [['niveau satisfaction'], ['satisfaction vis']])
        self.col_recommandation = find_col_any(df, [['recommanderiez-vous cette'], ['recommanderiez']])

        # Construire l'identifiant agence pour chaque ligne
        self.df['__agence__'] = self.df.apply(lambda r: get_agence_id(r, self.city_cols), axis=1)

    def _err(self, idx: int, rule: str, col: str, val_actuelle, correction: str, explication: str):
        row = self.df.iloc[idx]
        rname, severity = self.RULES_META.get(rule, (rule, 'mineure'))
        self.errors.append({
            'N° ligne données': idx + 2,
            'SbjNum': row.get(self.col_sbjnum, ''),
            'Enquêteur': row.get(self.col_srvyr, '') if self.col_srvyr else '',
            'Wilaya': row.get(self.col_wilaya, '') if self.col_wilaya else '',
            'Agence': row['__agence__'][:70],
            'Règle': rule,
            'Description règle': rname,
            'Colonne concernée': col,
            'Valeur actuelle': str(val_actuelle)[:100],
            'Correction suggérée': str(correction)[:150],
            'Explication': explication[:300],
            'Sévérité': severity,
        })

    # ----------------------------------------------------------
    # RÈGLE 1 : Couverture — 3 visites par agence
    # ----------------------------------------------------------
    def _r1_coverage(self):
        for agence, grp in self.df.groupby('__agence__'):
            if agence == 'AGENCE_INCONNUE':
                continue
            n = len(grp)
            if n != 3:
                msg = f"L'agence a {n} visite(s) au lieu de 3 requises."
                corr = 'Compléter à 3 visites' if n < 3 else 'Supprimer le(s) doublon(s)'
                for idx in grp.index:
                    self._err(idx, 'R1', 'Agence', f'{n} visite(s)', corr, msg)

    # ----------------------------------------------------------
    # RÈGLE 2 : Mix profils — 2 PRI + 1 PRO (ou inverse)
    # ----------------------------------------------------------
    def _r2_profile_mix(self):
        if not self.col_profil:
            return
        for agence, grp in self.df.groupby('__agence__'):
            if agence == 'AGENCE_INCONNUE' or len(grp) < 2:
                continue
            profiles = grp[self.col_profil].fillna('').str.upper().tolist()
            n_pri   = sum(1 for p in profiles if 'PRI' in p and 'PRO' not in p)
            n_pro   = sum(1 for p in profiles if 'PRO' in p and 'PRI' not in p)
            n_corpo = sum(1 for p in profiles if 'CORPO' in p)
            ok = (n_pri >= 2 and n_pro >= 1) or (n_pri >= 1 and n_pro >= 2)
            if not ok and len(grp) >= 3:
                detail = f"Profils trouvés : {n_pri} PRI, {n_pro} PRO, {n_corpo} CORPO. Requis : ≥2 PRI + ≥1 PRO ou ≥1 PRI + ≥2 PRO."
                for idx in grp.index:
                    val = self.df.at[idx, self.col_profil]
                    self._err(idx, 'R2', self.col_profil, val,
                              '2 PRI + 1 PRO ou 1 PRI + 2 PRO', detail)

    # ----------------------------------------------------------
    # RÈGLE 3 : Diversité scénarios — pas de doublon
    # ----------------------------------------------------------
    def _r3_scenario_diversity(self):
        checks = []
        if self.col_scen_pri:   checks.append(('PRI',   self.col_scen_pri))
        if self.col_scen_pro:   checks.append(('PRO',   self.col_scen_pro))
        if self.col_scen_corpo: checks.append(('CORPO', self.col_scen_corpo))

        for agence, grp in self.df.groupby('__agence__'):
            if agence == 'AGENCE_INCONNUE' or len(grp) < 2:
                continue
            for ptype, col in checks:
                vals = grp[col].dropna().astype(str).str.strip().tolist()
                vals = [v for v in vals if v and v.lower() != 'nan']
                if len(vals) >= 2 and len(set(vals)) < len(vals):
                    detail = f"Scénarios {ptype} en doublon pour cette agence : {', '.join(vals)}."
                    for idx in grp.index:
                        v = self.df.at[idx, col]
                        if pd.notna(v) and str(v).strip():
                            self._err(idx, 'R3', col, v,
                                      'Scénario différent des autres visites', detail)

    # ----------------------------------------------------------
    # RÈGLE 4 : Cohérence note / satisfaction / recommandation
    # ----------------------------------------------------------
    def _r4_satisfaction_coherence(self):
        if not self.col_note or not self.col_satisfaction:
            return
        for idx, row in self.df.iterrows():
            note_raw = row.get(self.col_note)
            sat_raw  = row.get(self.col_satisfaction)
            reco_raw = row.get(self.col_recommandation) if self.col_recommandation else None

            if pd.isna(note_raw) or pd.isna(sat_raw):
                continue
            try:
                note = float(note_raw)
            except (ValueError, TypeError):
                continue

            sat  = sat_to_score(sat_raw)
            reco = reco_to_score(reco_raw)

            # Note élevée (≥8) mais satisfaction négative
            if note >= 8 and sat in (1, 2):
                self._err(idx, 'R4', self.col_satisfaction,
                          f'Note={note}/10 | Satisfaction={sat_raw}',
                          'Satisfaction positive (Satisfait ou Très satisfait)',
                          f'Note valorisation {note}/10 incompatible avec satisfaction «{sat_raw}».')

            # Note basse (≤4) mais satisfaction positive
            elif note <= 4 and sat in (4, 5):
                self._err(idx, 'R4', self.col_satisfaction,
                          f'Note={note}/10 | Satisfaction={sat_raw}',
                          'Satisfaction négative (Insatisfait ou Très insatisfait)',
                          f'Note valorisation {note}/10 incompatible avec satisfaction «{sat_raw}».')

            # Recommandation incohérente avec satisfaction
            if reco != -1 and sat != -1:
                if sat >= 4 and reco <= 2:
                    self._err(idx, 'R4', self.col_recommandation,
                              f'Satisfaction={sat_raw} | Recommandation={reco_raw}',
                              'Recommandation positive si satisfaction élevée',
                              f'Satisfaction «{sat_raw}» mais recommandation négative «{reco_raw}».')
                elif sat <= 2 and reco >= 4:
                    self._err(idx, 'R4', self.col_recommandation,
                              f'Satisfaction={sat_raw} | Recommandation={reco_raw}',
                              'Recommandation négative si satisfaction basse',
                              f'Satisfaction «{sat_raw}» mais recommandation positive «{reco_raw}».')

    # ----------------------------------------------------------
    # RÈGLE 5 : Heure de visite ↔ créneau déclaré
    # ----------------------------------------------------------
    def _r5_time_coherence(self):
        if not self.col_heure or not self.col_heure_grille:
            return
        for idx, row in self.df.iterrows():
            heure  = row.get(self.col_heure)
            grille = row.get(self.col_heure_grille)
            if pd.isna(heure) or pd.isna(grille):
                continue
            slot = parse_time_slot(grille)
            if not slot:
                continue
            hs, ms, he, me = slot
            if hasattr(heure, 'hour'):
                ah, am = heure.hour, heure.minute
            else:
                continue
            actual = ah * 60 + am
            if not (hs * 60 + ms <= actual <= he * 60 + me):
                self._err(idx, 'R5', self.col_heure_grille,
                          f'Heure={ah:02d}h{am:02d} / Créneau={grille}',
                          f'Créneau {ah:02d}h00-{ah+1:02d}h00 (approx.)',
                          f'Heure réelle {ah:02d}h{am:02d} hors du créneau déclaré «{grille}».')

    # ----------------------------------------------------------
    # RÈGLE 6 : Temps d'attente ↔ créneau
    # ----------------------------------------------------------
    def _r6_wait_time(self):
        if not self.col_att_min or not self.col_att_grille:
            return
        for idx, row in self.df.iterrows():
            mins   = row.get(self.col_att_min)
            grille = row.get(self.col_att_grille)
            if pd.isna(mins) or pd.isna(grille):
                continue
            try:
                m = int(float(mins))
            except (ValueError, TypeError):
                continue
            slot = parse_minute_slot(grille)
            if not slot:
                continue
            lo, hi = slot
            if not (lo <= m <= hi):
                self._err(idx, 'R6', self.col_att_grille,
                          f'{m} min / Créneau={grille}',
                          f'Créneau correspondant à {m} minutes',
                          f"Temps d'attente ({m} min) hors du créneau déclaré «{grille}».")

    # ----------------------------------------------------------
    # RÈGLE 7 : Temps conseiller ↔ créneau
    # ----------------------------------------------------------
    def _r7_advisor_time(self):
        if not self.col_cons_min or not self.col_cons_grille:
            return
        for idx, row in self.df.iterrows():
            mins   = row.get(self.col_cons_min)
            grille = row.get(self.col_cons_grille)
            if pd.isna(mins) or pd.isna(grille):
                continue
            try:
                m = int(float(mins))
            except (ValueError, TypeError):
                continue
            slot = parse_minute_slot(grille)
            if not slot:
                continue
            lo, hi = slot
            if not (lo <= m <= hi):
                self._err(idx, 'R7', self.col_cons_grille,
                          f'{m} min / Créneau={grille}',
                          f'Créneau correspondant à {m} minutes',
                          f'Temps avec conseiller ({m} min) hors du créneau déclaré «{grille}».')

    # ----------------------------------------------------------
    # LANCEMENT COMPLET
    # ----------------------------------------------------------
    def run(self) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """
        Lance toutes les règles.
        Retourne (df_avec_flags, df_corrections)
        """
        self.errors = []
        self._r1_coverage()
        self._r2_profile_mix()
        self._r3_scenario_diversity()
        self._r4_satisfaction_coherence()
        self._r5_time_coherence()
        self._r6_wait_time()
        self._r7_advisor_time()

        # ---- Fichier 1 : données + colonnes de flags ----
        df_out = self.df.drop(columns=['__agence__'], errors='ignore').copy()

        # Index erreurs par SbjNum
        err_idx: Dict = {}
        for e in self.errors:
            k = e['SbjNum']
            err_idx.setdefault(k, []).append(e)

        nb_col   = []
        reg_col  = []
        det_col  = []
        col_sbjnum = self.col_sbjnum

        for _, row in df_out.iterrows():
            key = row.get(col_sbjnum, '')
            errs = err_idx.get(key, [])
            nb_col.append(len(errs))
            reg_col.append(' | '.join(sorted(set(e['Règle'] for e in errs))))
            det_col.append(' || '.join(e['Explication'] for e in errs))

        df_out.insert(0, 'Nb_problèmes',    nb_col)
        df_out.insert(1, 'Règles_violées',  reg_col)
        df_out.insert(2, 'Détail_problèmes', det_col)

        # ---- Fichier 2 : corrections ----
        if self.errors:
            df_corr = pd.DataFrame(self.errors)
            # Trier par sévérité puis règle
            sev_order = {'critique': 0, 'moderee': 1, 'mineure': 2}
            df_corr['_sev_ord'] = df_corr['Sévérité'].map(sev_order).fillna(9)
            df_corr = df_corr.sort_values(['_sev_ord', 'Règle', 'N° ligne données'])
            df_corr = df_corr.drop(columns=['_sev_ord'])
        else:
            df_corr = pd.DataFrame(columns=[
                'N° ligne données', 'SbjNum', 'Enquêteur', 'Wilaya', 'Agence',
                'Règle', 'Description règle', 'Colonne concernée',
                'Valeur actuelle', 'Correction suggérée', 'Explication', 'Sévérité'
            ])

        return df_out, df_corr

    def get_summary(self) -> Dict:
        """Résumé statistique des erreurs"""
        total = len(self.errors)
        by_rule = {}
        by_sev  = {'critique': 0, 'moderee': 0, 'mineure': 0}
        lignes_ok = set(self.df[self.col_sbjnum].tolist())
        lignes_ko = set()
        for e in self.errors:
            r = e['Règle']
            by_rule[r] = by_rule.get(r, 0) + 1
            by_sev[e['Sévérité']] = by_sev.get(e['Sévérité'], 0) + 1
            lignes_ko.add(e['SbjNum'])
        return {
            'total_erreurs': total,
            'lignes_avec_erreur': len(lignes_ko),
            'lignes_total': len(self.df),
            'by_rule': by_rule,
            'by_severite': by_sev,
        }


# ============================================================
# EXPORT EXCEL AVEC MISE EN FORME
# ============================================================

def _apply_header_style(ws, row=1):
    header_fill = PatternFill(fill_type='solid', fgColor='1F4E79')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    for cell in ws[row]:
        if cell.value is not None:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)


def export_to_excel(df_data: pd.DataFrame, df_corrections: pd.DataFrame) -> bytes:
    """
    Génère les 2 feuilles dans un seul fichier Excel avec mise en forme.
    Retourne les bytes du fichier.
    """
    output = io.BytesIO()

    ROUGE_CLAIR  = PatternFill(fill_type='solid', fgColor='FFCCCC')
    ORANGE_CLAIR = PatternFill(fill_type='solid', fgColor='FFE5CC')
    VERT_CLAIR   = PatternFill(fill_type='solid', fgColor='E2EFDA')

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # ---- Feuille 1 : Données + flags ----
        df_data.to_excel(writer, sheet_name='Données_vérifiées', index=False)
        ws1 = writer.sheets['Données_vérifiées']
        _apply_header_style(ws1)

        # Colorer les lignes selon nb_problèmes
        nb_col_idx = 1  # colonne A = Nb_problèmes
        for row_idx in range(2, ws1.max_row + 1):
            nb_cell = ws1.cell(row=row_idx, column=nb_col_idx)
            try:
                nb = int(nb_cell.value or 0)
            except (ValueError, TypeError):
                nb = 0

            if nb == 0:
                fill = VERT_CLAIR
            else:
                # Vérifier s'il y a des erreurs critiques
                rules_cell = ws1.cell(row=row_idx, column=2)
                rules_val = str(rules_cell.value or '')
                fill = ROUGE_CLAIR if 'R1' in rules_val or 'R2' in rules_val else ORANGE_CLAIR

            for col_idx in range(1, min(4, ws1.max_column + 1)):
                ws1.cell(row=row_idx, column=col_idx).fill = fill

        # Ajuster largeurs colonnes 1-3
        ws1.column_dimensions['A'].width = 14
        ws1.column_dimensions['B'].width = 20
        ws1.column_dimensions['C'].width = 60

        # Figer la première ligne
        ws1.freeze_panes = 'A2'

        # ---- Feuille 2 : Corrections ----
        df_corrections.to_excel(writer, sheet_name='Corrections_suggérées', index=False)
        ws2 = writer.sheets['Corrections_suggérées']
        _apply_header_style(ws2)

        # Colorer selon sévérité
        sev_col_idx = None
        for cidx, cell in enumerate(ws2[1], 1):
            if cell.value and 'v' in str(cell.value).lower() and 'rit' in str(cell.value).lower():
                sev_col_idx = cidx
                break
        # Fallback: chercher colonne "Sévérité"
        if not sev_col_idx:
            for cidx, cell in enumerate(ws2[1], 1):
                if cell.value and normalize(str(cell.value)) in ('severite', 'severité'):
                    sev_col_idx = cidx
                    break

        for row_idx in range(2, ws2.max_row + 1):
            sev = ''
            if sev_col_idx:
                sev = str(ws2.cell(row=row_idx, column=sev_col_idx).value or '').lower()
            if 'critique' in sev:
                fill = ROUGE_CLAIR
            elif 'moderee' in sev or 'modérée' in sev:
                fill = ORANGE_CLAIR
            else:
                fill = PatternFill(fill_type='solid', fgColor='FFFACC')
            for col_idx in range(1, ws2.max_column + 1):
                ws2.cell(row=row_idx, column=col_idx).fill = fill

        # Largeurs colonnes corrections
        col_widths = [12, 12, 14, 16, 40, 8, 35, 30, 30, 50, 60, 12]
        for i, w in enumerate(col_widths, 1):
            if i <= ws2.max_column:
                ws2.column_dimensions[get_column_letter(i)].width = w

        ws2.freeze_panes = 'A2'

    output.seek(0)
    return output.read()
